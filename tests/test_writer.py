"""
test_writer.py — Tests para reporting/writer.py (v2)
"""
import pytest
import pandas as pd
from unittest.mock import patch
from openpyxl import load_workbook

from reporting.writer import escribir_resultado, escribir_sin_conciliar


# ─── Fixture ──────────────────────────────────────────────────────────────────

@pytest.fixture
def df_resultado():
    return pd.DataFrame({
        # — Lado cartola —
        "fecha_operacion_cartola": pd.to_datetime(["2024-01-15", "2024-02-20", "2024-03-10"]),
        "fecha_valor_cartola":     pd.to_datetime(["2024-01-15", "2024-02-20", "2024-03-10"]),
        "glosa_cartola":           ["pago luz enel", "transferencia juan", "sueldo empresa"],
        "rut_cartola":             ["19141427-6", "21493875-8", "76354771-9"],
        "monto_cartola":           [-100_000.0, -50_000.0, 200_000.0],
        "nro_documento_cartola":   ["1234567890", "ABCD567890", "9999999999"],
        "banco_cartola":           ["Banco de Chile", "BancoEstado", "BCI"],
        # — Lado libro —
        "fecha_contable_libro":    pd.to_datetime(["2024-01-15", "2024-02-21", None]),
        "glosa_libro":             ["pago luz enel", "transferencia juan", None],
        "rut_libro":               ["19141427-6", "21493875-8", None],
        "monto_libro":             [-100_000.0, -50_200.0, None],
        "nro_referencia_libro":    ["1234567890", "XXXX567890", None],
        "nro_comprobante_libro":   ["CP001", "CP002", None],
        "codigo_tx_libro":         ["SRV001", "TRF002", None],
        # — Match —
        "tipo_match":              ["Exacto", "Sugerido", "Manual"],
        "certeza":                 ["Exacto", "Sugerido", "Manual"],
        "regla_aplicada":          ["RUT + Monto + Fecha", "RUT + Monto", ""],
        "diff_monto":              [0.0, 200.0, None],
        "diff_dias":               [0, 1, None],
        "flag_conciliacion":       ["", "", ""],
        "flag_iva":                ["", "", ""],
        # — Antigüedad —
        "dias_antiguedad":         [20, 60, 100],
        "tramo_antiguedad":        ["Vigente", "En Observación", "Crítico"],
        "accion_recomendada":      [
            "Aprobado — sin acción requerida",
            "Revisar y aprobar match manualmente",
            "Investigar y contabilizar manualmente",
        ],
        # — Diagnóstico (solo sin match) —
        "motivo":                  [None, None, "Monto coincide pero fecha fuera de rango"],
        "fecha_cercana":           pd.to_datetime([None, None, "2024-01-15"]),
        "monto_cercano":           [None, None, -100_000.0],
        "glosa_cercana":           [None, None, "pago luz enel"],
        "diff_monto_cercano":      [None, None, 300_000.0],
    })


# ─── escribir_resultado ───────────────────────────────────────────────────────

class TestEscribirResultado:

    def test_crea_archivo_xlsx(self, df_resultado, tmp_path):
        archivo = tmp_path / "resultado.xlsx"
        with patch("reporting.writer.ARCHIVO_RESULTADO", archivo), \
             patch("reporting.writer.OUTPUT_DIR", tmp_path):
            escribir_resultado(df_resultado)
        assert archivo.exists()

    def test_archivo_tiene_hoja_conciliacion(self, df_resultado, tmp_path):
        archivo = tmp_path / "resultado.xlsx"
        with patch("reporting.writer.ARCHIVO_RESULTADO", archivo), \
             patch("reporting.writer.OUTPUT_DIR", tmp_path):
            escribir_resultado(df_resultado)
        wb = load_workbook(archivo)
        assert "Conciliación" in wb.sheetnames

    def test_archivo_tiene_hoja_resumen(self, df_resultado, tmp_path):
        archivo = tmp_path / "resultado.xlsx"
        with patch("reporting.writer.ARCHIVO_RESULTADO", archivo), \
             patch("reporting.writer.OUTPUT_DIR", tmp_path):
            escribir_resultado(df_resultado)
        wb = load_workbook(archivo)
        assert "Resumen" in wb.sheetnames

    def test_fila_1_tiene_encabezados_agrupados(self, df_resultado, tmp_path):
        """La fila 1 debe tener los nombres de los bloques."""
        archivo = tmp_path / "resultado.xlsx"
        with patch("reporting.writer.ARCHIVO_RESULTADO", archivo), \
             patch("reporting.writer.OUTPUT_DIR", tmp_path):
            escribir_resultado(df_resultado)
        wb = load_workbook(archivo)
        ws = wb["Conciliación"]
        valores_fila1 = [ws.cell(row=1, column=c).value for c in [1, 8, 15]]
        assert "Cartola Personal" in valores_fila1
        assert "Libro del Banco"  in valores_fila1
        assert "Resultado"        in valores_fila1

    def test_fila_2_tiene_encabezados_columna(self, df_resultado, tmp_path):
        """Los encabezados de columna deben estar en la fila 2."""
        archivo = tmp_path / "resultado.xlsx"
        with patch("reporting.writer.ARCHIVO_RESULTADO", archivo), \
             patch("reporting.writer.OUTPUT_DIR", tmp_path):
            escribir_resultado(df_resultado)
        wb = load_workbook(archivo)
        ws = wb["Conciliación"]
        assert ws.cell(row=2, column=1).value == "Fecha Operación"
        assert ws.cell(row=2, column=15).value == "Tipo Match"

    def test_datos_empiezan_en_fila_3(self, df_resultado, tmp_path):
        """Los datos deben empezar en la fila 3."""
        archivo = tmp_path / "resultado.xlsx"
        with patch("reporting.writer.ARCHIVO_RESULTADO", archivo), \
             patch("reporting.writer.OUTPUT_DIR", tmp_path):
            escribir_resultado(df_resultado)
        wb = load_workbook(archivo)
        ws = wb["Conciliación"]
        assert ws.max_row == len(df_resultado) + 2

    def test_cantidad_columnas_correcta(self, df_resultado, tmp_path):
        """7 cartola + 7 libro + 10 resultado = 24 columnas."""
        archivo = tmp_path / "resultado.xlsx"
        with patch("reporting.writer.ARCHIVO_RESULTADO", archivo), \
             patch("reporting.writer.OUTPUT_DIR", tmp_path):
            escribir_resultado(df_resultado)
        wb = load_workbook(archivo)
        ws = wb["Conciliación"]
        assert ws.max_column == 24


# ─── escribir_sin_conciliar ───────────────────────────────────────────────────

class TestEscribirSinConciliar:

    def test_crea_archivo_xlsx(self, df_resultado, tmp_path):
        archivo = tmp_path / "sin_conciliar.xlsx"
        with patch("reporting.writer.ARCHIVO_SIN_CONCILIAR", archivo), \
             patch("reporting.writer.OUTPUT_DIR", tmp_path):
            escribir_sin_conciliar(df_resultado)
        assert archivo.exists()

    def test_tiene_hoja_sin_conciliar(self, df_resultado, tmp_path):
        archivo = tmp_path / "sin_conciliar.xlsx"
        with patch("reporting.writer.ARCHIVO_SIN_CONCILIAR", archivo), \
             patch("reporting.writer.OUTPUT_DIR", tmp_path):
            escribir_sin_conciliar(df_resultado)
        wb = load_workbook(archivo)
        assert "Sin Conciliar" in wb.sheetnames

    def test_solo_contiene_filas_sin_match(self, df_resultado, tmp_path):
        """1 bloque + 1 encabezado + 1 fila Manual = 3 filas."""
        archivo = tmp_path / "sin_conciliar.xlsx"
        with patch("reporting.writer.ARCHIVO_SIN_CONCILIAR", archivo), \
             patch("reporting.writer.OUTPUT_DIR", tmp_path):
            escribir_sin_conciliar(df_resultado)
        wb = load_workbook(archivo)
        ws = wb["Sin Conciliar"]
        assert ws.max_row == 3   # fila bloque + fila encabezado + 1 dato

    def test_tiene_10_columnas(self, df_resultado, tmp_path):
        """7 cartola + 3 diagnóstico = 10 columnas."""
        archivo = tmp_path / "sin_conciliar.xlsx"
        with patch("reporting.writer.ARCHIVO_SIN_CONCILIAR", archivo), \
             patch("reporting.writer.OUTPUT_DIR", tmp_path):
            escribir_sin_conciliar(df_resultado)
        wb = load_workbook(archivo)
        ws = wb["Sin Conciliar"]
        assert ws.max_column == 10

    def test_fila_1_tiene_bloques_correctos(self, df_resultado, tmp_path):
        archivo = tmp_path / "sin_conciliar.xlsx"
        with patch("reporting.writer.ARCHIVO_SIN_CONCILIAR", archivo), \
             patch("reporting.writer.OUTPUT_DIR", tmp_path):
            escribir_sin_conciliar(df_resultado)
        wb = load_workbook(archivo)
        ws = wb["Sin Conciliar"]
        valores = [ws.cell(row=1, column=c).value for c in [1, 8]]
        assert "Cartola Personal" in valores
        assert "Diagnóstico"      in valores

    def test_encabezado_motivo_presente(self, df_resultado, tmp_path):
        """La columna Motivo debe estar en el encabezado."""
        archivo = tmp_path / "sin_conciliar.xlsx"
        with patch("reporting.writer.ARCHIVO_SIN_CONCILIAR", archivo), \
             patch("reporting.writer.OUTPUT_DIR", tmp_path):
            escribir_sin_conciliar(df_resultado)
        wb = load_workbook(archivo)
        ws = wb["Sin Conciliar"]
        encabezados_fila2 = [ws.cell(row=2, column=c).value for c in range(1, 11)]
        assert "Motivo" in encabezados_fila2


# ─── Verificación de archivo bloqueado ───────────────────────────────────────

class TestArchivoBloqueado:

    def test_error_claro_si_resultado_esta_abierto(self, df_resultado, tmp_path):
        """Si el archivo está bloqueado, debe lanzar ConciliadorError."""
        from utils.exceptions import ConciliadorError
        from unittest.mock import patch as mock_patch

        archivo = tmp_path / "resultado.xlsx"
        archivo.touch()

        with mock_patch("reporting.writer.ARCHIVO_RESULTADO", archivo), \
             mock_patch("reporting.writer.OUTPUT_DIR", tmp_path), \
             mock_patch("pathlib.Path.unlink", side_effect=PermissionError):
            with pytest.raises(ConciliadorError, match="está abierto en Excel"):
                escribir_resultado(df_resultado)

    def test_error_claro_si_sin_conciliar_esta_abierto(self, df_resultado, tmp_path):
        """Si el archivo sin conciliar está bloqueado, debe lanzar ConciliadorError."""
        from utils.exceptions import ConciliadorError
        from unittest.mock import patch as mock_patch

        archivo = tmp_path / "sin_conciliar.xlsx"
        archivo.touch()

        with mock_patch("reporting.writer.ARCHIVO_SIN_CONCILIAR", archivo), \
             mock_patch("reporting.writer.OUTPUT_DIR", tmp_path), \
             mock_patch("pathlib.Path.unlink", side_effect=PermissionError):
            with pytest.raises(ConciliadorError, match="está abierto en Excel"):
                escribir_sin_conciliar(df_resultado)


# ─── Hoja Resumen ─────────────────────────────────────────────────────────────

class TestHojaResumen:

    def _cargar_resumen(self, df_resultado, tmp_path):
        archivo = tmp_path / "resultado.xlsx"
        with patch("reporting.writer.ARCHIVO_RESULTADO", archivo), \
             patch("reporting.writer.OUTPUT_DIR", tmp_path):
            escribir_resultado(df_resultado)
        return load_workbook(archivo)["Resumen"]

    def test_total_transacciones_correcto(self, df_resultado, tmp_path):
        ws = self._cargar_resumen(df_resultado, tmp_path)
        assert ws.cell(row=3, column=2).value == 3

    def test_exactos_correcto(self, df_resultado, tmp_path):
        ws = self._cargar_resumen(df_resultado, tmp_path)
        assert ws.cell(row=4, column=2).value == 1

    def test_sugeridos_correcto(self, df_resultado, tmp_path):
        ws = self._cargar_resumen(df_resultado, tmp_path)
        assert ws.cell(row=5, column=2).value == 1

    def test_manuales_correcto(self, df_resultado, tmp_path):
        ws = self._cargar_resumen(df_resultado, tmp_path)
        assert ws.cell(row=6, column=2).value == 1

    def test_porcentaje_conciliado(self, df_resultado, tmp_path):
        ws = self._cargar_resumen(df_resultado, tmp_path)
        assert ws.cell(row=7, column=2).value == "66.7%"