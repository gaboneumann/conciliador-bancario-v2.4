"""
formatter.py — Estilos y colores para los Excel de salida (v2.2).

Responsabilidad única: definir y retornar objetos de estilo de openpyxl.
No escribe archivos ni manipula datos.

Colores de filas por tipo de match (v2):
    Exacto           → verde claro        #C6EFCE
    Sugerido         → amarillo claro     #FFEB9C
    Manual           → rojo claro         #FFC7CE
    flag_conciliacion→ azul claro         #BDD7EE  (prioridad sobre Sugerido)
    flag_iva         → verde agua         #E2EFDA

Colores de encabezados agrupados:
    Cartola Personal → azul oscuro        #1F4E79
    Libro del Banco  → verde oscuro       #1A5632
    Resultado        → gris oscuro        #404040
    Diagnóstico      → rojo oscuro        #7B2C2C

Colores hoja Hallazgos_Criticos (v2.2):
    Concentración >20% → rojo claro      #FFC7CE
    Crítico >90 días   → naranja claro   #FFCC99
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─── Colores de filas ─────────────────────────────────────────────────────────

COLORES = {
    # v2
    "Exacto":        "C6EFCE",   # verde claro
    "Sugerido":      "FFEB9C",   # amarillo claro
    "Manual":        "FFC7CE",   # rojo claro
    "conciliacion":  "BDD7EE",   # azul claro  (flag Partida en Conciliación)
    "iva":           "E2EFDA",   # verde agua  (flag IVA)
    # v2.2 — Hallazgos_Criticos
    "concentracion": "FFC7CE",   # rojo claro  (RUT concentra >20% del error)
    "critico":       "FFCC99",   # naranja claro (antigüedad >90 días)
    # v1 — mantenidos por retrocompatibilidad
    "exacto":        "C6EFCE",
    "parcial":       "FFEB9C",
    "sin_match":     "FFC7CE",
    # utilidades
    "encabezado":    "1F4E79",   # azul oscuro (encabezado simple)
    "blanco":        "FFFFFF",
    "gris_claro":    "F2F2F2",
}

# ─── Colores de bloques (encabezados agrupados) ───────────────────────────────

COLORES_BLOQUE = {
    "cartola":      "1F4E79",  # azul oscuro
    "libro":        "1A5632",  # verde oscuro
    "resultado":    "404040",  # gris oscuro
    "diagnostico":  "7B2C2C",  # rojo oscuro
    "hallazgos":    "7B2C2C",  # rojo oscuro (mismo que diagnóstico)
}

# ─── Fuente ───────────────────────────────────────────────────────────────────

FUENTE  = "Arial"
TAMANO  = 10


# ─── Constructores de estilo ──────────────────────────────────────────────────

def estilo_encabezado() -> dict:
    """Estilo para la fila de encabezados de columna (fila 2)."""
    return {
        "font": Font(
            name=FUENTE,
            size=TAMANO,
            bold=True,
            color="FFFFFF",
        ),
        "fill": PatternFill(
            fill_type="solid",
            start_color=COLORES["encabezado"],
        ),
        "alignment": Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        ),
        "border": _borde_fino(),
    }


def estilo_encabezado_bloque(bloque: str) -> dict:
    """
    Estilo para la fila superior de encabezados agrupados (fila 1).

    Args:
        bloque: "cartola", "libro", "resultado", "diagnostico" o "hallazgos"
    """
    color = COLORES_BLOQUE.get(bloque, COLORES_BLOQUE["resultado"])

    return {
        "font": Font(
            name=FUENTE,
            size=TAMANO + 1,
            bold=True,
            color="FFFFFF",
        ),
        "fill": PatternFill(
            fill_type="solid",
            start_color=color,
        ),
        "alignment": Alignment(
            horizontal="center",
            vertical="center",
        ),
        "border": _borde_medio(),
    }


def estilo_fila(tipo_match: str, flag_conciliacion: str = "", flag_iva: str = "") -> dict:
    """
    Estilo para una fila de datos según su tipo de match y flags activos.

    Prioridad de color:
        1. flag_conciliacion activo → azul claro  (#BDD7EE)
        2. flag_iva activo          → verde agua  (#E2EFDA)
        3. tipo_match               → Exacto / Sugerido / Manual
    """
    if flag_conciliacion:
        color = COLORES["conciliacion"]
    elif flag_iva:
        color = COLORES["iva"]
    else:
        color = COLORES.get(tipo_match, COLORES["blanco"])

    return {
        "font":      Font(name=FUENTE, size=TAMANO),
        "fill":      PatternFill(fill_type="solid", start_color=color),
        "alignment": Alignment(vertical="center"),
        "border":    _borde_fino(),
    }


def estilo_hallazgo(alerta: str, tramo: str) -> dict:
    """
    Estilo para filas de la hoja Hallazgos_Criticos.

    Prioridad de color:
        1. alerta concentración >20% → rojo   (#FFC7CE)
        2. tramo Crítico >90 días    → naranja (#FFCC99)
        3. resto                     → blanco

    Args:
        alerta: string con "⚠️" si hay concentración, "" si no
        tramo : "Crítico", "En Observación" o "Vigente"
    """
    if alerta:
        color = COLORES["concentracion"]
    elif tramo == "Crítico":
        color = COLORES["critico"]
    else:
        color = COLORES["blanco"]

    return {
        "font":      Font(name=FUENTE, size=TAMANO),
        "fill":      PatternFill(fill_type="solid", start_color=color),
        "alignment": Alignment(vertical="center"),
        "border":    _borde_fino(),
    }

def estilo_texto_naranja() -> dict:
    """Estilo de texto naranja para partidas Críticas en hallazgos."""
    return {
        "font": Font(name=FUENTE, size=TAMANO, color="FF6600"),
        "fill": PatternFill(fill_type="solid", start_color=COLORES["blanco"]),
        "alignment": Alignment(vertical="center"),
        "border": _borde_fino(),
    }

def estilo_numero() -> dict:
    """Estilo adicional para celdas con montos."""
    return {
        "number_format": '#,##0_-;[Red](#,##0)',
        "alignment":     Alignment(horizontal="right", vertical="center"),
    }


def estilo_fecha() -> dict:
    """Estilo adicional para celdas con fechas."""
    return {
        "number_format": "YYYY-MM-DD",
        "alignment":     Alignment(horizontal="center", vertical="center"),
    }


# ─── Anchos de columna v2 ─────────────────────────────────────────────────────

ANCHOS_RESULTADO = {
    "A": 14, "B": 14, "C": 35, "D": 18, "E": 16, "F": 18, "G": 20,
    "H": 14, "I": 35, "J": 18, "K": 16, "L": 18, "M": 16, "N": 14,
    "O": 12, "P": 12, "Q": 28, "R": 14, "S": 12,
    "T": 26, "U": 26, "V": 14, "W": 18, "X": 45,
}

ANCHOS_SIN_CONCILIAR = {
    "A": 14, "B": 14, "C": 35, "D": 18, "E": 16, "F": 18, "G": 20,
    "H": 45, "I": 16, "J": 16,
}

ANCHOS_HALLAZGOS = {
    "A": 18,   # RUT
    "B": 38,   # Glosa Frecuente
    "C": 18,   # Cantidad Partidas
    "D": 22,   # Monto Total Pendiente
    "E": 28,   # Motivo Principal
    "F": 20,   # Antigüedad Máxima (días)
    "G": 20,   # % sobre Total Error
    "H": 28,   # Alerta
}

# ─── Definición de bloques ────────────────────────────────────────────────────

BLOQUES_RESULTADO = [
    {"nombre": "Cartola Personal", "bloque": "cartola",   "col_inicio": 1,  "col_fin": 7},
    {"nombre": "Libro del Banco",  "bloque": "libro",     "col_inicio": 8,  "col_fin": 14},
    {"nombre": "Resultado",        "bloque": "resultado", "col_inicio": 15, "col_fin": 24},
]

BLOQUES_SIN_CONCILIAR = [
    {"nombre": "Cartola Personal", "bloque": "cartola",     "col_inicio": 1, "col_fin": 7},
    {"nombre": "Diagnóstico",      "bloque": "diagnostico", "col_inicio": 8, "col_fin": 10},
]

BLOQUES_HALLAZGOS = [
    {"nombre": "Hallazgos Críticos — Ranking por RUT", "bloque": "hallazgos", "col_inicio": 1, "col_fin": 8},
]


# ─── Utilidades internas ──────────────────────────────────────────────────────

def _borde_fino() -> Border:
    lado = Side(style="thin", color="BFBFBF")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _borde_medio() -> Border:
    lado_grueso = Side(style="medium", color="FFFFFF")
    lado_fino   = Side(style="thin",   color="BFBFBF")
    return Border(
        left=lado_grueso, right=lado_grueso,
        top=lado_fino,    bottom=lado_fino,
    )