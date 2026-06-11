"""
writer.py — Escritura de archivos Excel de resultado (v2.3f).

CAMBIOS v2.3f:
─────────────────────────────────────────────────────────────────────────────
- IDs Referencia: columna concatenada con nro_documento (cartola/sugerido)
  y nro_comprobante (libro sin par), separados por " | ".
─────────────────────────────────────────────────────────────────────────────
"""
import pandas as pd
from datetime import date
from pathlib import Path
from openpyxl import Workbook

from config.config import (
    ARCHIVO_RESULTADO,
    ARCHIVO_SIN_CONCILIAR,
    ARCHIVO_HALLAZGOS,
    OUTPUT_DIR,
)
from reporting.formatter import (
    estilo_encabezado,
    estilo_encabezado_bloque,
    estilo_fila,
    estilo_numero,
    estilo_fecha,
    ANCHOS_RESULTADO,
    ANCHOS_SIN_CONCILIAR,
    BLOQUES_RESULTADO,
    BLOQUES_SIN_CONCILIAR,
)
from utils.logger import get_logger
from utils.exceptions import ConciliadorError

logger = get_logger(__name__)

UMBRAL_CONCENTRACION = 0.20
HOY = pd.Timestamp(date.today())

BLOQUES_HALLAZGOS = [
    {
        "nombre":     "Hallazgos Críticos — Ranking por RUT",
        "bloque":     "hallazgos",
        "col_inicio": 1,
        "col_fin":    9,
    },
]

ANCHOS_HALLAZGOS = {
    "A": 18,
    "B": 38,
    "C": 18,
    "D": 24,
    "E": 22,
    "F": 45,
    "G": 22,
    "H": 22,
    "I": 40,
}

ENCABEZADOS_HALLAZGOS = [
    "RUT",
    "Glosa Frecuente",
    "Cant. Partidas",
    "Monto de Impacto",
    "Motivo Principal",
    "IDs Referencia",
    "Antigüedad Máxima (días)",
    "% sobre Error",
    "Plan de Acción",
]

# ─── Resolver rutas de output ─────────────────────────────────────────────────

def _resolver_rutas(output_dir: "Path | None" = None) -> tuple:
    """
    Retorna (archivo_resultado, archivo_sin_conciliar, archivo_hallazgos, carpeta).
    Si output_dir es None, usa las rutas de config.py (modo consola).
    Si output_dir viene con valor, construye las rutas sobre esa carpeta (modo GUI).
    """
    if output_dir is None:
        return ARCHIVO_RESULTADO, ARCHIVO_SIN_CONCILIAR, ARCHIVO_HALLAZGOS, OUTPUT_DIR

    carpeta = Path(output_dir)
    return (
        carpeta / ARCHIVO_RESULTADO.name,
        carpeta / ARCHIVO_SIN_CONCILIAR.name,
        carpeta / ARCHIVO_HALLAZGOS.name,
        carpeta,
    )

# ─── Verificación de archivo disponible ──────────────────────────────────────

def _verificar_archivo_disponible(ruta) -> None:
    if ruta.exists():
        try:
            ruta.unlink()
        except PermissionError:
            raise ConciliadorError(
                f"El archivo '{ruta.name}' está abierto en Excel. "
                f"Ciérralo e intenta de nuevo."
            )

# ─── Definición de columnas ──────────────────────────────────────────────────

ENCABEZADOS_RESULTADO = [
    "Fecha Operación", "Fecha Valor", "Glosa Cartola",
    "RUT Cartola", "Monto Cartola", "Nº Documento", "Banco",
    "Fecha Contable", "Glosa Libro",
    "RUT Libro", "Monto Libro", "Nº Referencia", "Nº Comprobante", "Código Tx",
    "Tipo Match", "Certeza", "Regla Aplicada",
    "Dif. Monto", "Dif. Días",
    "Flag Conciliación", "Flag IVA",
    "Días Antigüedad", "Tramo Antigüedad", "Acción Recomendada",
]

COLUMNAS_RESULTADO = [
    "fecha_operacion_cartola", "fecha_valor_cartola", "glosa_cartola",
    "rut_cartola", "monto_cartola", "nro_documento_cartola", "banco_cartola",
    "fecha_contable_libro", "glosa_libro",
    "rut_libro", "monto_libro", "nro_referencia_libro",
    "nro_comprobante_libro", "codigo_tx_libro",
    "tipo_match", "certeza", "regla_aplicada",
    "diff_monto", "diff_dias",
    "flag_conciliacion", "flag_iva",
    "dias_antiguedad", "tramo_antiguedad", "accion_recomendada",
]

ENCABEZADOS_SIN_CONCILIAR = [
    "Fecha Operación", "Fecha Valor", "Glosa Cartola",
    "RUT Cartola", "Monto Cartola", "Nº Documento", "Banco",
    "Motivo", "Monto Más Cercano", "Dif. Monto Cercano",
]

COLUMNAS_SIN_CONCILIAR = [
    "fecha_operacion_cartola", "fecha_valor_cartola", "glosa_cartola",
    "rut_cartola", "monto_cartola", "nro_documento_cartola", "banco_cartola",
    "motivo", "monto_cercano", "diff_monto_cercano",
]

COLS_MONTO = {
    "monto_cartola", "monto_libro", "diff_monto",
    "monto_cercano", "diff_monto_cercano",
}
COLS_FECHA = {
    "fecha_operacion_cartola", "fecha_valor_cartola",
    "fecha_contable_libro",
}

# ─── Fila de encabezados agrupados ───────────────────────────────────────────

def _escribir_fila_bloques(ws, bloques: list) -> None:
    ws.row_dimensions[1].height = 22
    for bloque in bloques:
        col_ini = bloque["col_inicio"]
        col_fin = bloque["col_fin"]
        if col_ini < col_fin:
            ws.merge_cells(
                start_row=1, start_column=col_ini,
                end_row=1,   end_column=col_fin,
            )
        celda           = ws.cell(row=1, column=col_ini, value=bloque["nombre"])
        estilo          = estilo_encabezado_bloque(bloque["bloque"])
        celda.font      = estilo["font"]
        celda.fill      = estilo["fill"]
        celda.alignment = estilo["alignment"]
        celda.border    = estilo["border"]

# ─── Escritura de hoja estándar ───────────────────────────────────────────────

def _escribir_hoja(ws, df, columnas, encabezados, anchos, bloques) -> None:
    ws.sheet_view.showGridLines = False
    _escribir_fila_bloques(ws, bloques)

    ws.row_dimensions[2].height = 36
    estilo_enc = estilo_encabezado()
    for col_idx, nombre in enumerate(encabezados, start=1):
        celda           = ws.cell(row=2, column=col_idx, value=nombre)
        celda.font      = estilo_enc["font"]
        celda.fill      = estilo_enc["fill"]
        celda.alignment = estilo_enc["alignment"]
        celda.border    = estilo_enc["border"]

    for row_idx, (_, fila) in enumerate(df[columnas].iterrows(), start=3):
        tipo_match        = fila.get("tipo_match", "Manual")
        flag_conciliacion = fila.get("flag_conciliacion", "")
        estilo            = estilo_fila(tipo_match, flag_conciliacion=flag_conciliacion)

        for col_idx, col_nombre in enumerate(columnas, start=1):
            valor = fila[col_nombre]
            if pd.isna(valor):
                valor = None
            celda           = ws.cell(row=row_idx, column=col_idx, value=valor)
            celda.font      = estilo["font"]
            celda.fill      = estilo["fill"]
            celda.alignment = estilo["alignment"]
            celda.border    = estilo["border"]
            if col_nombre in COLS_MONTO and valor is not None:
                celda.number_format = estilo_numero()["number_format"]
                celda.alignment     = estilo_numero()["alignment"]
            elif col_nombre in COLS_FECHA and valor is not None:
                celda.number_format = estilo_fecha()["number_format"]
                celda.alignment     = estilo_fecha()["alignment"]

    for letra, ancho in anchos.items():
        ws.column_dimensions[letra].width = ancho
    ws.freeze_panes = "A3"

# ─── Construir DataFrame de hallazgos (v2.3f) ────────────────────────────────

def _construir_hallazgos(
    df_resultado: pd.DataFrame,
    df_libro: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """
    Tres familias de MI para cuadrar con diferencia de saldo:

        1. Manual        → MI = monto_cartola       | ID = nro_documento_cartola
        2. Sugerido      → MI = monto_c - monto_l   | ID = nro_documento_cartola
        3. Libro sin Par → MI = -monto_libro        | ID = nro_comprobante

    sum(MI) == saldo_cartola - saldo_libro ✅
    IDs Referencia: concatenados con " | " por RUT para trazabilidad de auditoría.
    """
    mapa_motivo = {
        "Fecha coincide pero monto no encontrado":  "Omisión",
        "Monto coincide pero fecha fuera de rango": "Corte",
        "Posible Neto vs Bruto (×1.19)":            "IVA",
        "Transacción ausente en libro auxiliar":    "Ausente",
    }

    partes = []

    # — Familia 1: Manuales —
    manuales = df_resultado[df_resultado["tipo_match"] == "Manual"].copy()
    if not manuales.empty:
        manuales["_mi"]     = manuales["monto_cartola"]
        manuales["_motivo"] = manuales.get("motivo", pd.Series("", index=manuales.index))
        manuales["_rut"]    = manuales["rut_cartola"].fillna("RUT NO IDENTIFICADO")
        manuales["_glosa"]  = manuales["glosa_cartola"]
        manuales["_fecha"]  = pd.to_datetime(manuales["fecha_valor_cartola"], errors="coerce")
        manuales["_id"]     = manuales["nro_documento_cartola"].fillna("S/N").astype(str)
        partes.append(manuales[["_rut", "_glosa", "_mi", "_motivo", "_fecha", "_id"]])

    # — Familia 2: Sugeridos con MI != 0 —
    sugeridos = df_resultado[df_resultado["tipo_match"] == "Sugerido"].copy()
    if not sugeridos.empty:
        sugeridos["_mi"] = sugeridos["monto_cartola"] - sugeridos["monto_libro"].fillna(0)
        sugeridos = sugeridos[sugeridos["_mi"].abs() > 0]
        if not sugeridos.empty:
            sugeridos["_motivo"] = sugeridos["flag_iva"].apply(
                lambda x: "IVA" if x else "Materialidad"
            )
            sugeridos["_rut"]   = sugeridos["rut_cartola"].fillna("RUT NO IDENTIFICADO")
            sugeridos["_glosa"] = sugeridos["glosa_cartola"]
            sugeridos["_fecha"] = pd.to_datetime(sugeridos["fecha_valor_cartola"], errors="coerce")
            sugeridos["_id"]    = sugeridos["nro_documento_cartola"].fillna("S/N").astype(str)
            partes.append(sugeridos[["_rut", "_glosa", "_mi", "_motivo", "_fecha", "_id"]])

    # — Familia 3: Libro sin Par —
    if df_libro is not None and "idx_libro" in df_resultado.columns:
        indices_matcheados = set(
            df_resultado["idx_libro"]
            .dropna()
            .astype(int)
        )
        libro_sin_par = df_libro.loc[
            ~df_libro.index.isin(indices_matcheados)
        ].copy()

        if not libro_sin_par.empty:
            libro_sin_par["_mi"]     = -libro_sin_par["monto"]
            libro_sin_par["_motivo"] = "Libro sin Par"
            libro_sin_par["_rut"]    = libro_sin_par["rut"].fillna("RUT NO IDENTIFICADO") if "rut" in libro_sin_par.columns else "RUT NO IDENTIFICADO"
            libro_sin_par["_glosa"]  = libro_sin_par["glosa"] if "glosa" in libro_sin_par.columns else ""
            libro_sin_par["_fecha"]  = pd.to_datetime(
                libro_sin_par["fecha_contable"] if "fecha_contable" in libro_sin_par.columns else None,
                errors="coerce"
            )
            libro_sin_par["_id"] = (
                libro_sin_par["nro_comprobante"].fillna("S/N").astype(str)
                if "nro_comprobante" in libro_sin_par.columns else "S/N"
            )
            partes.append(libro_sin_par[["_rut", "_glosa", "_mi", "_motivo", "_fecha", "_id"]])
            logger.info(
                f"Libro sin par: {len(libro_sin_par)} filas → "
                f"MI = {libro_sin_par['_mi'].sum():,.0f}"
            )

    if not partes:
        return pd.DataFrame()

    combinado = pd.concat(partes, ignore_index=True)
    combinado["_rut"] = combinado["_rut"].fillna("RUT NO IDENTIFICADO")
    combinado.loc[
        combinado["_rut"].astype(str).str.strip() == "", "_rut"
    ] = "RUT NO IDENTIFICADO"
    combinado["_antiguedad"] = (
        HOY - combinado["_fecha"]
    ).dt.days.fillna(0).astype(int)

    diferencia_total = combinado["_mi"].sum()

    filas = []
    for rut, grupo in combinado.groupby("_rut"):
        glosa_frecuente = (
            grupo["_glosa"].dropna().mode().iloc[0]
            if not grupo["_glosa"].dropna().empty else ""
        )
        motivo_counts    = grupo["_motivo"].value_counts()
        motivo_raw       = motivo_counts.index[0] if not motivo_counts.empty else "Sin diagnóstico"
        motivo_principal = mapa_motivo.get(motivo_raw, motivo_raw)

        cantidad       = len(grupo)
        mi_rut         = grupo["_mi"].sum()
        antiguedad_max = grupo["_antiguedad"].max()
        pct_error      = (
            abs(mi_rut) / abs(diferencia_total) * 100
            if diferencia_total != 0 else 0
        )
        alerta = "⚠️ Riesgo de Concentración Alto" if pct_error > UMBRAL_CONCENTRACION * 100 else ""
        ids    = " | ".join(grupo["_id"].dropna().unique().tolist())

        filas.append({
            "rut":             rut,
            "glosa_frecuente": glosa_frecuente,
            "cantidad":        cantidad,
            "monto_impacto":   round(mi_rut, 0),
            "motivo":          motivo_principal,
            "ids_referencia":  ids,
            "antiguedad_max":  antiguedad_max,
            "pct_error":       round(pct_error, 1),
            "plan_accion":     "",
            "alerta":          alerta,
        })

    df_h = pd.DataFrame(filas)
    df_h["_es_sin_rut"] = (df_h["rut"] == "RUT NO IDENTIFICADO").astype(int)
    df_h = (
        df_h.sort_values(["_es_sin_rut", "antiguedad_max"], ascending=[False, False])
        .drop(columns=["_es_sin_rut"])
        .reset_index(drop=True)
    )
    return df_h

# ─── Escritura de hoja hallazgos ─────────────────────────────────────────────

def _escribir_hoja_hallazgos(ws, df_hallazgos: pd.DataFrame) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    from reporting.formatter import _borde_fino

    FUENTE = "Arial"
    TAMANO = 10

    ws.sheet_view.showGridLines = False
    _escribir_fila_bloques(ws, BLOQUES_HALLAZGOS)

    ws.row_dimensions[2].height = 36
    estilo_enc = estilo_encabezado()
    for col_idx, nombre in enumerate(ENCABEZADOS_HALLAZGOS, start=1):
        celda           = ws.cell(row=2, column=col_idx, value=nombre)
        celda.font      = estilo_enc["font"]
        celda.fill      = estilo_enc["fill"]
        celda.alignment = estilo_enc["alignment"]
        celda.border    = estilo_enc["border"]

    if df_hallazgos.empty:
        ws.cell(row=3, column=1, value="Sin hallazgos — conciliación completa ✅")
        return

    columnas_df = [
        "rut", "glosa_frecuente", "cantidad", "monto_impacto",
        "motivo", "ids_referencia", "antiguedad_max", "pct_error", "plan_accion",
    ]

    borde = _borde_fino()

    for row_idx, (_, fila) in enumerate(df_hallazgos.iterrows(), start=3):
        alerta     = fila["alerta"]
        antiguedad = fila["antiguedad_max"]
        es_critico = antiguedad > 90
        es_sin_rut = fila["rut"] == "RUT NO IDENTIFICADO"

        if alerta:
            fill       = PatternFill(fill_type="solid", start_color="FFC7CE")
            font_color = "000000"
        elif es_critico:
            fill       = PatternFill(fill_type="solid", start_color="FFFFFF")
            font_color = "9C5600"
        else:
            fill       = PatternFill(fill_type="solid", start_color="FFFFFF")
            font_color = "000000"

        for col_idx, col_nombre in enumerate(columnas_df, start=1):
            valor = fila[col_nombre]
            if pd.isna(valor):
                valor = None
            celda           = ws.cell(row=row_idx, column=col_idx, value=valor)
            celda.font      = Font(name=FUENTE, size=TAMANO, color=font_color, bold=es_sin_rut)
            celda.fill      = fill
            celda.alignment = Alignment(
                vertical="center",
                wrap_text=(col_nombre == "ids_referencia"),
            )
            celda.border    = borde

            if col_nombre == "monto_impacto" and valor is not None:
                celda.number_format = estilo_numero()["number_format"]
                celda.alignment     = estilo_numero()["alignment"]
            elif col_nombre == "pct_error" and valor is not None:
                celda.number_format = '0.0"%"'
                celda.alignment     = estilo_numero()["alignment"]
            elif col_nombre in ("cantidad", "antiguedad_max") and valor is not None:
                celda.alignment = estilo_numero()["alignment"]

    for letra, ancho in ANCHOS_HALLAZGOS.items():
        ws.column_dimensions[letra].width = ancho
    ws.freeze_panes = "A3"

# ─── Funciones públicas ───────────────────────────────────────────────────────

def escribir_resultado(
    df_resultado: pd.DataFrame,
    saldo: dict | None = None,
    output_dir: "Path | None" = None,
) -> None:
    archivo_resultado, _, _, carpeta = _resolver_rutas(output_dir)
    carpeta.mkdir(parents=True, exist_ok=True)
    _verificar_archivo_disponible(archivo_resultado)

    wb            = Workbook()
    ws_conc       = wb.active
    ws_conc.title = "Conciliación"
    _escribir_hoja(
        ws_conc, df_resultado,
        COLUMNAS_RESULTADO, ENCABEZADOS_RESULTADO,
        ANCHOS_RESULTADO, BLOQUES_RESULTADO,
    )
    ws_resumen = wb.create_sheet("Resumen")
    _escribir_resumen(ws_resumen, df_resultado, saldo=saldo)

    wb.save(archivo_resultado)
    logger.info(f"Resultado guardado → {archivo_resultado}")


def escribir_sin_conciliar(
    df_resultado: pd.DataFrame,
    output_dir: "Path | None" = None,
) -> None:
    from conciliation.classifier import separar_sin_conciliar

    _, archivo_sin_conciliar, _, carpeta = _resolver_rutas(output_dir)
    carpeta.mkdir(parents=True, exist_ok=True)
    _verificar_archivo_disponible(archivo_sin_conciliar)

    df_sin   = separar_sin_conciliar(df_resultado)
    wb       = Workbook()
    ws       = wb.active
    ws.title = "Sin Conciliar"
    _escribir_hoja(
        ws, df_sin,
        COLUMNAS_SIN_CONCILIAR, ENCABEZADOS_SIN_CONCILIAR,
        ANCHOS_SIN_CONCILIAR, BLOQUES_SIN_CONCILIAR,
    )
    wb.save(archivo_sin_conciliar)
    logger.info(
        f"Sin conciliar guardado → {archivo_sin_conciliar} ({len(df_sin)} partidas)"
    )


def escribir_hallazgos(
    df_resultado: pd.DataFrame,
    saldo: dict | None = None,
    df_libro: pd.DataFrame | None = None,
    output_dir: "Path | None" = None,
) -> None:
    _, _, archivo_hallazgos, carpeta = _resolver_rutas(output_dir)
    carpeta.mkdir(parents=True, exist_ok=True)
    _verificar_archivo_disponible(archivo_hallazgos)

    df_hallazgos = _construir_hallazgos(df_resultado, df_libro)

    if df_hallazgos.empty:
        logger.info("Sin hallazgos — archivo no generado ✅")
        return

    if saldo:
        suma     = df_hallazgos["monto_impacto"].sum()
        esperado = saldo["diferencia"]
        delta    = abs(suma - esperado)
        if delta < 1:
            logger.info(f"Hallazgos cuadra con diferencia de saldo ✅ ({suma:,.0f})")
        else:
            logger.warning(
                f"Hallazgos NO cuadra ⚠️ "
                f"(hallazgos={suma:,.0f} | saldo={esperado:,.0f} | delta={delta:,.0f})"
            )

    wb       = Workbook()
    ws       = wb.active
    ws.title = "Hallazgos_Criticos"
    _escribir_hoja_hallazgos(ws, df_hallazgos)

    wb.save(archivo_hallazgos)
    logger.info(
        f"Hallazgos guardado → {archivo_hallazgos} ({len(df_hallazgos)} RUTs)"
    )


def _escribir_resumen(ws, df: pd.DataFrame, saldo: dict | None) -> None:
    from reporting.formatter import _borde_fino
    ws.sheet_view.showGridLines = False

    total     = len(df)
    exactos   = (df["tipo_match"] == "Exacto").sum()
    sugeridos = (df["tipo_match"] == "Sugerido").sum()
    manuales  = (df["tipo_match"] == "Manual").sum()
    pct_conc  = round((exactos + sugeridos) / total * 100, 1) if total > 0 else 0

    filas = [
        ("Resumen de Conciliación", ""),
        ("", ""),
        ("Total transacciones cartola", total),
        ("Match Exacto",               exactos),
        ("Match Sugerido",             sugeridos),
        ("Sin Match (Manual)",         manuales),
        ("% Conciliado",               f"{pct_conc}%"),
    ]

    if saldo:
        filas += [
            ("", ""),
            ("Diferencia de Saldo", ""),
            ("Saldo Cartola",  saldo["saldo_cartola"]),
            ("Saldo Libro",    saldo["saldo_libro"]),
            ("Diferencia",     saldo["diferencia"]),
            ("¿Cuadra?",       "Sí" if saldo["cuadra"] else "No"),
        ]

    estilo_enc  = estilo_encabezado()
    estilo_dato = estilo_fila("blanco")
    borde       = _borde_fino()

    for row_idx, (etiqueta, valor) in enumerate(filas, start=1):
        celda_a = ws.cell(row=row_idx, column=1, value=etiqueta)
        celda_b = ws.cell(row=row_idx, column=2, value=valor)
        es_titulo = row_idx == 1 or etiqueta in ("Diferencia de Saldo",)

        if es_titulo and etiqueta:
            celda_a.font = estilo_enc["font"]
            celda_a.fill = estilo_enc["fill"]
            celda_b.fill = estilo_enc["fill"]
        else:
            celda_a.font = estilo_dato["font"]
            celda_b.font = estilo_dato["font"]

        celda_a.border = borde
        celda_b.border = borde

        if isinstance(valor, (int, float)) and etiqueta in (
            "Saldo Cartola", "Saldo Libro", "Diferencia"
        ):
            celda_b.number_format = estilo_numero()["number_format"]

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18